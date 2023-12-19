"""This module is responsible for reading debitor data from emails in Outlook."""

from io import BytesIO
import json

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from itk_dev_shared_components.graph import authentication, mail
from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection

from robot_framework import config
from robot_framework.sub_process import common


def load_debitors_from_emails(orchestrator_connection: OrchestratorConnection):
    """Load debitor data from all the emails in
    "itk-rpa@mkb.aarhus.dk" - "Indbakke/Statstidende/Debitor Udtræk".

    Returns:
        A set of unique debitors in the format (fp, id, name, street, street_no, zip code)
    """
    orchestrator_connection.log_info("Fetching data from OPUS emails")

    graph_creds = orchestrator_connection.get_credential(config.GRAPH_API)
    graph_access = authentication.authorize_by_username_password(graph_creds.username, **json.loads(graph_creds.password))

    emails = mail.get_emails_from_folder("itk-rpa@mkb.aarhus.dk", "Indbakke/Statstidende/Debitor Udtræk", graph_access)

    debitors = set()

    for email in emails:
        orchestrator_connection.log_info(f"Reading Email: {email.subject}")
        att = mail.list_email_attachments(email, graph_access)[0]
        excel_file = mail.get_attachment_data(att, graph_access)
        read_sheet(excel_file, debitors)
        excel_file.close()

    orchestrator_connection.log_info(f"Debitore i Opus: {len(debitors)}")
    return debitors


def delete_emails(orchestrator_connection: OrchestratorConnection):
    """Delete the emails with debitors from Outlook.

    Args:
        orchestrator_connection: The connection to Orchestrator.
    """
    orchestrator_connection.log_info("Deleting OPUS emails")

    graph_creds = orchestrator_connection.get_credential(config.GRAPH_API)
    graph_access = authentication.authorize_by_username_password(graph_creds.username, **json.loads(graph_creds.password))

    emails = mail.get_emails_from_folder("itk-rpa@mkb.aarhus.dk", "Indbakke/Statstidende/Debitor Udtræk", graph_access)
    for email in emails:
        orchestrator_connection.log_info(f"Deleting Email: {email.subject}")
        mail.delete_email(email, graph_access)


def read_sheet(excel_file: BytesIO, debitors: set[str]) -> None:
    """Read an Excel sheet and adds debitors to the given set.

    Args:
        excel_file: The excel file to read.
        debitors: The set to add debitors to.
    """
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    ws = wb.active

    rows = ws.values
    column_names = next(rows)
    aftaletype_index = column_names.index('RIM aftaletype')

    for row in rows:
        # Skip rows with aftaletype 'IN'
        if row[aftaletype_index] != 'IN':
            # Remove aftaletype column before adding to set
            row = tuple(e for index, e in enumerate(row) if index != aftaletype_index)
            debitors.add(row)

    wb.close()


def find_relevant_cases(in_cases: list[dict], orchestrator_connection: OrchestratorConnection) -> list[tuple[str]]:
    """Find all Statstidende cases that could have relevance for debitors in OPUS.

    Args:
        in_cases: A list of Statstidende cases.

    Returns:
        A list of relevant cases.
    """
    dødsboer, gældssaneringer, konkursboer, tvangsauktioner = in_cases
    debitors = load_debitors_from_emails(orchestrator_connection)

    out_cases = []

    for debitor in debitors:
        debitor_id = debitor[1]
        first_name = debitor[2].split()[0]
        street = debitor[4]
        # husnummer = row[5] Not used
        zipcode = debitor[6]
        birthdate = common.get_birthdate(debitor_id)

        # Search on cpr
        if debitor_id in dødsboer:
            out_cases.append((debitor, dødsboer[debitor_id]))

        # Search on cvr
        if debitor_id in konkursboer:
            out_cases.append((debitor, konkursboer[debitor_id]))

        # Search on birthdate and first name
        if birthdate in gældssaneringer:
            for case in gældssaneringer[birthdate]:
                if first_name in case[0]:
                    out_cases.append((debitor, case))

        # Search on street and zipcode
        if street and zipcode:
            for address in tvangsauktioner:
                if common.compare_addresses(address, street, zipcode):
                    out_cases.append((debitor, tvangsauktioner[address]))

    return out_cases


def write_excel(path: str, cases: tuple[tuple[str]]):
    """Write the given cases to an excel file on the given path.

    Args:
        path: Where to save the excel file.
        cases: A tuple of cases in the format:
            ((Debitor), (Case)) = ((fp, id, fornavn, efternavn, gade, husnr, postnummer, by), (X, Type, Sagsnummer, dato))
    """
    wb = openpyxl.Workbook()

    # Create sheets
    doedsboer_sheet = wb.active
    doedsboer_sheet.title = "Dødsboer"
    gaeldssaneringer_sheet = wb.create_sheet("Gældssaneringer")
    konkursboer_sheet = wb.create_sheet("Konkursboer")
    tvangsauktioner_sheet = wb.create_sheet("Tvangsauktioner")

    # Add column names to sheets
    doedsboer_sheet.append(("CPR", "Navn", "Adresse", "CPR på Sag", "Type", "Sagsnummer", "Sagsdato"))
    gaeldssaneringer_sheet.append(("CPR", "Navn", "Adresse", "Navn på sag", "Type", "Sagsnummer", "Sagsdato"))
    konkursboer_sheet.append(("CVR", "Navn", "Adresse", "CVR på sag", "Type", "Sagsnummer", "Sagsdato"))
    tvangsauktioner_sheet.append(("ID", "Navn", "Adresse", "Adresse på sag", "Type", "Sagsnummer", "Sagsdato"))

    for case in cases:
        case_type = case[1][1]

        debitor_id = case[0][1]
        name = " ".join(filter(None, case[0][2:4]))
        address = " ".join(filter(None, case[0][4:8]))
        data = [debitor_id, name, address] + case[1]

        if case_type.startswith("Dødsboer"):
            doedsboer_sheet.append(data)
        elif case_type.startswith("Gældssanering"):
            gaeldssaneringer_sheet.append(data)
        elif case_type.startswith("Konkursboer"):
            konkursboer_sheet.append(data)
        elif case_type.startswith("Tvangsauktioner"):
            tvangsauktioner_sheet.append(data)

    # Create excel tables for better formatting
    for ws in wb:
        dim = ws.dimensions
        if dim == 'A1:G1':
            continue
        table = Table(displayName=ws.title, ref=dim)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        ws.add_table(table)

    wb.save(path)
    wb.close()
