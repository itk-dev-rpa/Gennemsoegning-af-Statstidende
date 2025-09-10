"""This module is responsible for all logic concerning KMD Boliglån"""

import csv
import os
import time
import re
import subprocess

from _ctypes import COMError
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
import uiautomation
from itk_dev_shared_components.misc import file_util

from robot_framework.sub_process import common


def login(username: str, password: str):
    """Launch and login to KMD Boliglån."""
    subprocess.Popen(r"C:\Program Files (x86)\KMD\KMD.LW.Boliglaan\KMD.LW.KMDBoliglaan.Client.exe")  # pylint: disable=consider-using-with

    kmd_logon = uiautomation.WindowControl(AutomationId="MainLogonWindow", searchDepth=1)

    # Wait for logon window to load
    for _ in range(5):
        if len(kmd_logon.ComboBoxControl(AutomationId="UserPwComboBoxCics").GetSelectionPattern().GetSelection()) == 1:
            break
        time.sleep(1)

    kmd_logon.EditControl(AutomationId="UserPwTextBoxUserName").GetValuePattern().SetValue(username)
    kmd_logon.EditControl(AutomationId="UserPwPasswordBoxPassword").GetValuePattern().SetValue(password)
    kmd_logon.ButtonControl(AutomationId="UserPwLogonButton").GetInvokePattern().Invoke()

    boliglaan = uiautomation.WindowControl(Name="KMD Boliglån", searchDepth=1)
    if not boliglaan.Exists(maxSearchSeconds=20):
        raise RuntimeError("Boliglån didn't appear within 20 seconds")


def load_lenders() -> list[tuple[str]]:
    """Go through KMD Boliglån and save a list of lenders based on filter
    criteria. Read the list and return the data.
    """
    laanestatus = [
        "Bevilget",
        "Udbetalt",
        "Afvikling - frivillig",
        "Afvikling - 5 års løbetid"
    ]

    laanetype = [
        "Pligtlån 5 års løbetid §56, §57",
        "Pligtlån sanering §54",
        "Pligtlån pensionister §54",
        "Pligtlån enkeltværelse §54",
        "Pligtlån genhusning §54",
        "Pligtlån ældrebolig §54",
        "Flygtningelån 100% 5 års løbetid §66, §67",
        "Flygtningelån 100% sanering §66, §67",
        "Flygtningelån 100% pensionister §66, §67",
        "Flygtninglån 50% 5 års løbetid §68",
        "Kommune lån §59"
    ]

    # Open search window
    boliglaan = uiautomation.WindowControl(Name="KMD Boliglån", searchDepth=1)
    boliglaan.SendKeys("{Ctrl}b")

    # Check boxes and search
    laanesager_window = boliglaan.WindowControl(Name="Udsøg lånesager", searchDepth=1)

    for control, _ in uiautomation.WalkControl(laanesager_window):
        if isinstance(control, uiautomation.CheckBoxControl) and control.Name in laanestatus+laanetype:
            control.GetTogglePattern().Toggle(waitTime=0)

    laanesager_window.ButtonControl(Name="Søg").GetInvokePattern().Invoke()

    # Wait for search
    save_button = boliglaan.GroupControl(AutomationId="SagerLayoutPanel", searchDepth=4).ToolBarControl(AutomationId="Bar").ButtonControl(Name="Save")
    for _ in range(5):
        try:
            save_button.GetInvokePattern().Invoke()
            break
        except COMError:
            pass
    else:
        raise TimeoutError("Boliglån result didn't appear after a long time")

    # Save file and read it
    folder = os.getcwd()
    path = os.path.join(os.getcwd(), "udtræk.csv")
    file_util.handle_save_dialog(path)
    file_util.wait_for_download(folder=folder, file_name="udtræk", file_extension=".csv")

    return read_csv(path)


def read_csv(file_name: str) -> list[tuple[str]]:
    """Read a csv file from KMD Boliglån and extract
    cpr, name and address for each lender.
    """
    lenders = []
    with open(file_name, encoding='ansi') as file:
        # Skip first 12 rows (Meta date)
        for _ in range(12):
            next(file)

        # Read data
        reader = csv.reader(file, delimiter=';')
        for row in reader:
            lender = (
                row[1],  # cpr
                row[2],  # Name
                row[5]   # Address
            )
            lenders.append(lender)

    if len(lenders) == 0:
        raise RuntimeError("Found no lenders from KMD Boliglån")

    return lenders


def find_relevant_cases(in_cases: list[dict], orchestrator_connection: OrchestratorConnection) -> list[tuple[str]]:
    """Find all Statstidende cases that could have relevance for lenders in KMD Boliglån.

    Args:
        in_cases: A list of Statstidende cases.

    Returns:
        A list of relevant cases.
    """
    doedsboer, gaeldssaneringer, _, tvangsauktioner = in_cases

    orchestrator_connection.log_info("Finder lånere i Boliglån.")
    lenders = load_lenders()
    orchestrator_connection.log_info(f"Lånere i boliglån: {len(lenders)}")

    out_cases = []

    for lender in lenders:
        cpr = lender[0]
        name = lender[1]
        first_name = name.split()[0]
        address = lender[2]
        street = address.split()[0]
        zipcode = get_zipcode(address)
        birthdate = common.get_birthdate(cpr)

        # Search dødsboer on cpr
        if cpr in doedsboer:
            out_cases.append((lender, doedsboer[cpr]))

        # Search gældssaneringer on birthdate and first name
        if birthdate in gaeldssaneringer:
            for case in gaeldssaneringer[birthdate]:
                if first_name in case[0]:
                    out_cases.append((lender, case))

        # Search tvangsauktioner on street and zipcode
        if street and zipcode:
            for address in tvangsauktioner:
                if common.compare_addresses(address, street, zipcode):
                    out_cases.append((lender, tvangsauktioner[address]))

    return out_cases


def get_zipcode(address: str) -> str:
    """Extract the zipcode from an address string.

    Args:
        address: The full address.

    Returns:
        The zipcode from the address. If none is found '9999' is returned.
    """
    # Use regular expressions to find all four-digit numbers
    pattern = r"\b\d{4}\b"
    matches = re.findall(pattern, address)

    if matches:
        postal_code = matches[-1]  # Get the last occurrence
        return postal_code

    return '9999'


# pylint: disable=R0801
def write_excel(path: str, cases: tuple[tuple[str]]) -> None:
    """Write the given cases to an excel file on the given path.

    Args:
        path: Where to save the excel file.
        cases: A tuple of cases in the format:
            ((Debitor), (Case)) = ((fp, id, fornavn, efternavn, gade, husnr, postnummer, by), (X, Type, Sagsnummer, dato))
    """
    # Cases = ((Debitor),(Case)) = ((cpr, navn, adresse),(X, Type, Sagsnummer, dato))
    wb = openpyxl.Workbook()
    doedsboer_sheet = wb.active
    doedsboer_sheet.title = "Dødsboer"
    gaeldssaneringer_sheet = wb.create_sheet("Gældssaneringer")
    tvangsauktioner_sheet = wb.create_sheet("Tvangsauktioner")

    doedsboer_sheet.append(("CPR", "Navn", "Adresse", "CPR på Sag", "Type", "Sagsnummer", "Sagsdato"))
    gaeldssaneringer_sheet.append(("CPR", "Navn", "Adresse", "Navn på sag", "Type", "Sagsnummer", "Sagsdato"))
    tvangsauktioner_sheet.append(("CPR", "Navn", "Adresse", "Adresse på sag", "Type", "Sagsnummer", "Sagsdato"))

    for case in cases:
        case_type = case[1][1]
        data = list(case[0]) + list(case[1])

        if case_type.startswith("Dødsboer"):
            doedsboer_sheet.append(data)
        elif case_type.startswith("Gældssanering"):
            gaeldssaneringer_sheet.append(data)
        elif case_type.startswith("Tvangsauktioner"):
            tvangsauktioner_sheet.append(data)

    for ws in wb:
        dim = ws.dimensions
        if dim == 'A1:G1':
            continue
        table = Table(displayName=ws.title, ref=dim)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        ws.add_table(table)

    wb.save(path)
    wb.close()


def kill_boliglaan():
    """Kill KMD Logon, KMD Boliglån and Notepad."""
    os.system("taskkill /f /im KMD.LW.KMDBoliglaan.Client.exe")
    os.system("taskkill /f /im KMD.YH.Security.Logon.Desktop.exe")
    os.system("taskkill /f /im notepad*")


if __name__ == '__main__':
    print(len(load_lenders()))
